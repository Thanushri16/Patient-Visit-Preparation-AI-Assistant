"""Load benchmark rows from Excel and turn them into executable conversations."""

from dataclasses import asdict, dataclass, replace
from pathlib import Path
import re

from openpyxl import load_workbook

try:
    from .preconditions import resolve_scenario_messages
except ImportError:  # pragma: no cover - allows running as a script
    from preconditions import resolve_scenario_messages


@dataclass(frozen=True)
class BenchmarkScenario:
    """One scored spreadsheet scenario, including any parsed conversation turns."""

    test_id: str
    category: str
    subcategory: str
    turn: str
    user_message: str
    expected_intent: str
    expected_behavior: str
    pass_fail_criteria: str
    tests_concept: str
    is_multi_turn: bool
    messages: tuple[str, ...] = ()
    # Turns that establish the state the row says already exists. They are sent
    # to the same session before `messages` and are never scored.
    setup_messages: tuple[str, ...] = ()

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


_TURN_MARKER = re.compile(
    r"(?:^|\s)Turn\s+(?:\d+(?:\s*[-–]\s*\d+)?):\s*",
    flags=re.IGNORECASE,
)


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_inline_turns(value: str) -> tuple[str, ...]:
    """Split `Turn N:` scripts while accepting quoted and unquoted messages."""

    matches = list(_TURN_MARKER.finditer(value))
    if not matches:
        return ()

    messages: list[str] = []
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(value)
        message = value[match.end():end].strip().strip("'\" ")
        if message:
            messages.append(message)
    return tuple(messages)


def _is_multi_turn(turn: str, user_message: str) -> bool:
    normalized = turn.strip()
    return normalized not in {"", "1", "1.0"} or len(parse_inline_turns(user_message)) > 1


def load_scenarios(path: Path) -> list[BenchmarkScenario]:
    """Read and validate all rows in the workbook's `Scenarios` sheet."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Scenarios" not in workbook.sheetnames:
        raise ValueError("Workbook must contain a 'Scenarios' sheet.")

    worksheet = workbook["Scenarios"]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [_text(value) for value in next(rows)]
    except StopIteration as exc:
        raise ValueError("The 'Scenarios' sheet is empty.") from exc

    required = [
        "Test ID",
        "Category",
        "Subcategory",
        "Turn",
        "User Message",
        "Expected Intent",
        "Expected Behavior",
        "Pass / Fail Criteria",
        "Tests Concept",
    ]
    missing_headers = [name for name in required if name not in headers]
    if missing_headers:
        raise ValueError(f"Missing workbook columns: {', '.join(missing_headers)}")

    scenarios: list[BenchmarkScenario] = []
    for values in rows:
        record = dict(zip(headers, values, strict=False))
        test_id = _text(record.get("Test ID"))
        if not test_id:
            continue
        user_message = _text(record.get("User Message"))
        turn = _text(record.get("Turn"))
        parsed_messages = parse_inline_turns(user_message)
        is_multi_turn = _is_multi_turn(turn, user_message)
        scenarios.append(
            BenchmarkScenario(
                test_id=test_id,
                category=_text(record.get("Category")),
                subcategory=_text(record.get("Subcategory")),
                turn=turn,
                user_message=user_message,
                expected_intent=_text(record.get("Expected Intent")),
                expected_behavior=_text(record.get("Expected Behavior")),
                pass_fail_criteria=_text(record.get("Pass / Fail Criteria")),
                tests_concept=_text(record.get("Tests Concept")),
                is_multi_turn=is_multi_turn,
                messages=parsed_messages or (user_message,),
            )
        )

    # Numeric continuation rows can be grouped when the workbook provides a
    # preceding row with the same category and subcategory. Inline scripts stay
    # self-contained and therefore require no neighboring-row assumptions.
    grouped: list[BenchmarkScenario] = []
    for scenario in scenarios:
        if (
            scenario.turn in {"2", "3", "4", "5"}
            and len(scenario.messages) == 1
            and grouped
            and grouped[-1].category == scenario.category
            and grouped[-1].subcategory == scenario.subcategory
        ):
            previous = grouped.pop()
            grouped.append(
                replace(
                    scenario,
                    messages=previous.messages + scenario.messages,
                    is_multi_turn=True,
                )
            )
        else:
            grouped.append(scenario)

    # Finally, resolve any stated precondition into the setup turns it
    # describes, so a row that says "(Summary shown)" is actually shown one.
    resolved: list[BenchmarkScenario] = []
    for scenario in grouped:
        setup, scored = resolve_scenario_messages(scenario.messages)
        resolved.append(replace(scenario, setup_messages=setup, messages=scored))
    return resolved


def split_scenarios(
    scenarios: list[BenchmarkScenario],
) -> tuple[list[BenchmarkScenario], list[BenchmarkScenario]]:
    """Return single-turn scenarios and multi-turn sequences separately."""

    singles = [scenario for scenario in scenarios if not scenario.is_multi_turn]
    multi_turn = [scenario for scenario in scenarios if scenario.is_multi_turn]
    return singles, multi_turn
