"""Load multi-turn conversation flows from the benchmark workbook.

The `Conversation Flows` sheet is one row per turn, grouped by `Conv ID`. Unlike
the `Scenarios` sheet, every turn of a flow belongs to the same session and the
order matters, so the loader keeps the turns together and preserves their order
rather than treating each row as an independent case.
"""

from dataclasses import asdict, dataclass, field
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_COLUMNS = (
    "Conv ID",
    "Flow Category",
    "Flow Name",
    "Turn #",
    "Speaker",
    "Message / Expected Check",
    "Expected Intent / State",
)


@dataclass(frozen=True)
class FlowTurn:
    """One user turn and the state the workbook expects it to produce."""

    number: int
    speaker: str
    message: str
    expectation: str
    notes: str = ""

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class ConversationFlow:
    """A whole session: ordered turns sharing one conversation state."""

    conv_id: str
    category: str
    name: str
    turns: tuple[FlowTurn, ...] = field(default_factory=tuple)

    @property
    def notes(self) -> str:
        """The flow-level intent, which the workbook records on its first turn."""

        return self.turns[0].notes if self.turns else ""

    def to_dict(self) -> dict[str, object]:
        return {
            "conv_id": self.conv_id,
            "category": self.category,
            "name": self.name,
            "notes": self.notes,
            "turns": [turn.to_dict() for turn in self.turns],
        }


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def load_conversation_flows(path: Path) -> list[ConversationFlow]:
    """Read and group every turn in the workbook's `Conversation Flows` sheet."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Conversation Flows" not in workbook.sheetnames:
        raise ValueError("Workbook must contain a 'Conversation Flows' sheet.")

    worksheet = workbook["Conversation Flows"]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [_text(value) for value in next(rows)]
    except StopIteration as exc:
        raise ValueError("The 'Conversation Flows' sheet is empty.") from exc

    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        raise ValueError(f"Missing workbook columns: {', '.join(missing)}")

    grouped: dict[str, list[FlowTurn]] = {}
    metadata: dict[str, tuple[str, str]] = {}
    order: list[str] = []

    for values in rows:
        record = dict(zip(headers, values, strict=False))
        conv_id = _text(record.get("Conv ID"))
        message = _text(record.get("Message / Expected Check"))
        if not conv_id or not message:
            continue
        if conv_id not in grouped:
            grouped[conv_id] = []
            order.append(conv_id)
            metadata[conv_id] = (
                _text(record.get("Flow Category")),
                _text(record.get("Flow Name")),
            )
        raw_turn = _text(record.get("Turn #"))
        grouped[conv_id].append(
            FlowTurn(
                number=int(float(raw_turn)) if raw_turn else len(grouped[conv_id]) + 1,
                speaker=_text(record.get("Speaker")) or "user",
                message=message,
                expectation=_text(record.get("Expected Intent / State")),
                notes=_text(record.get("Notes")),
            )
        )

    flows: list[ConversationFlow] = []
    for conv_id in order:
        category, name = metadata[conv_id]
        turns = tuple(sorted(grouped[conv_id], key=lambda turn: turn.number))
        flows.append(
            ConversationFlow(conv_id=conv_id, category=category, name=name, turns=turns)
        )
    return flows
