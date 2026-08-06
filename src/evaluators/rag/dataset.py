"""Loader for the RAG benchmark workbook."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_WORKBOOK = (
    Path(__file__).resolve().parents[1] / "rag_benchmark_questions.xlsx"
)
SHEET = "RAG Questions"


def _split(value) -> tuple[str, ...]:
    if not value:
        return ()
    return tuple(part.strip() for part in str(value).split(";") if part.strip())


@dataclass(frozen=True)
class BenchmarkCase:
    """One expectation about how the assistant should answer."""

    question_id: str
    group: str
    question: str
    expected_outcome: str
    expected_document_ids: tuple[str, ...] = ()
    expected_facts: tuple[str, ...] = ()
    expected_covered_facts: tuple[str, ...] = ()
    expected_uncovered_topics: tuple[str, ...] = ()
    forbidden_claims: tuple[str, ...] = ()
    near_miss_of: tuple[str, ...] = ()
    notes: str = ""

    @property
    def should_answer(self) -> bool:
        return self.expected_outcome in {"answered", "partially_answered"}

    @property
    def split(self) -> str:
        """Which half of the benchmark this case belongs to.

        Assigned by a hash of the question id, so the split is stable across
        runs and machines without storing it anywhere — a case cannot drift
        between halves when the workbook is edited or reordered.

        The split exists because tuning and reporting on the same questions
        makes the score optimistic by an unknown amount. Thresholds, the
        answerability prompt and the generation prompt were all tuned against
        the whole set; from here, tuning uses `tune` and the number quoted comes
        from `holdout`.
        """

        digest = hashlib.sha256(self.question_id.encode()).hexdigest()
        return "tune" if int(digest[:8], 16) % 2 == 0 else "holdout"


def load_cases(path: Path | None = None) -> list[BenchmarkCase]:
    sheet = load_workbook(path or DEFAULT_WORKBOOK, data_only=True)[SHEET]
    header = [cell.value for cell in sheet[1]]
    cases: list[BenchmarkCase] = []
    for row in sheet.iter_rows(min_row=2):
        record = dict(zip(header, [cell.value for cell in row]))
        if not record.get("question_id"):
            continue
        cases.append(
            BenchmarkCase(
                question_id=str(record["question_id"]),
                group=str(record.get("group") or ""),
                question=str(record["question"]),
                expected_outcome=str(record.get("expected_outcome") or ""),
                expected_document_ids=_split(record.get("expected_document_ids")),
                expected_facts=_split(record.get("expected_facts")),
                expected_covered_facts=_split(record.get("expected_covered_facts")),
                expected_uncovered_topics=_split(record.get("expected_uncovered_topics")),
                forbidden_claims=_split(record.get("forbidden_claims")),
                near_miss_of=_split(record.get("near_miss_of")),
                notes=str(record.get("notes") or ""),
            )
        )
    return cases
